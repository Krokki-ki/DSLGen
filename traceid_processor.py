#!/usr/bin/env python3
from __future__ import annotations

import math
import sys
import time
import re
from pathlib import Path
from typing import Iterable, Iterator, Tuple, List, Optional

try:
    from openpyxl import load_workbook
except ImportError:
    print("Для чтения XLSX требуется пакет 'openpyxl'. Установите: pip install openpyxl")
    sys.exit(1)

def write_lines(file_path: Path, lines: Iterable[str], newline: bool = True) -> None:
    file_path.parent.mkdir(parents=True, exist_ok=True)
    with file_path.open("w", encoding="utf-8") as f:
        if newline:
            for line in lines:
                f.write(line)
                if not line.endswith("\n"):
                    f.write("\n")
        else:
            for line in lines:
                f.write(line)

def progress_bar(current: int, total: int, prefix: str = "", width: int = 30) -> None:
    pct = 0.0 if total <= 0 else min(100.0, (current / total) * 100.0)
    filled = int(width * pct / 100)
    bar = "#" * filled + "." * (width - filled)
    sys.stdout.write(f"\r{prefix}[{bar}] {pct:5.1f}%")
    sys.stdout.flush()

def ask_inputs() -> Tuple[Path, str, Path]:
    directory_str = input("Введите путь к директории: ").strip()
    directory = Path(directory_str).expanduser().resolve()
    if not directory.is_dir():
        print("Указанная директория не существует.")
        sys.exit(1)

    filename_no_ext = input("Введите имя файла без расширения (XLSX): ").strip()
    if not filename_no_ext:
        print("Имя файла не должно быть пустым.")
        sys.exit(1)

    filename = filename_no_ext + ".xlsx"
    source_file = directory / filename
    if not source_file.is_file():
        print(f"Файл не найден: {source_file}")
        sys.exit(1)

    return directory, filename_no_ext, source_file

def normalize_trace_id(s: str) -> str:
    # Удаляем все виды пробельных символов внутри (пробелы, табы, \r, \n и т.п.)
    return re.sub(r"\s+", "", s)

def read_trace_ids_from_xlsx(path: Path, sheet_name: Optional[str] = None, col_index: Optional[int] = None) -> Iterator[str]:
    """
    Читает traceId из XLSX.
    - sheet_name: если None — активный лист.
    - col_index: если None — берётся первая непустая ячейка в строке; иначе 0-based индекс колонки (0=A, 1=B,...).
    Каждое значение приводится к строке, триммится и отдаётся наружу (без нормализации здесь).
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    for row in ws.iter_rows(values_only=True):  # быстрее с values_only [8][6]
        if not row:
            continue
        value = None
        if col_index is not None and 0 <= col_index < len(row):
            value = row[col_index]
        else:
            # Ищем первую непустую ячейку в строке
            for cell in row:
                if cell is not None and str(cell).strip() != "":
                    value = cell
                    break
        if value is None:
            continue
        s = str(value).strip()
        if s:
            yield s

def unique_trace_ids_with_progress_xlsx(file_path: Path) -> Tuple[List[str], int]:
    print("Выполняется поиск дублей записей (XLSX)...")
    # Для XLSX заранее оценить total_rows сложно без двойного прохода;
    # используем плавающий прогресс по счётчику
    seen: set[str] = set()
    out: List[str] = []
    duplicates = 0
    idx = 0
    # Читаем все значения по строкам
    for raw in read_trace_ids_from_xlsx(file_path):
        idx += 1
        if idx % 1000 == 0:
            progress_bar(idx, max(idx, 1), prefix="Поиск дублей: ")
        s = normalize_trace_id(raw)
        if not s:
            continue
        if s in seen:
            duplicates += 1
        else:
            seen.add(s)
            out.append(s)
    # финальный прогресс
    progress_bar(idx, max(idx, 1), prefix="Поиск дублей: ")
    sys.stdout.write("\n")
    if duplicates == 0:
        print("Массив подготовлен. Дублей не обнаружено")
    else:
        print(f"Массив подготовлен. Было найдено {duplicates} дублей.")
    return out, duplicates

def object_dsl_lines_for(tid: str) -> List[str]:
    # Один объект DSL — 5 строк
    return [
        "{",
        '  "match_phrase": {',
        f'    "traceId": "{tid}"',
        "  }",
        "},",
    ]

PREFIX = """{
  "query": {
    "bool": {
      "should": [
"""
SUFFIX = """  ],
      "minimum_should_match": 1
    }
  }
}
"""

def write_group_file_wrapped(target: Path, object_lines: List[str]) -> None:
    # Удаляем финальную запятую у последнего объекта
    if object_lines:
        for i in range(len(object_lines) - 1, -1, -1):
            if object_lines[i].strip() == "},":
                object_lines[i] = "}"
                break
    out_lines: List[str] = []
    out_lines.extend(PREFIX.splitlines())
    out_lines.extend(object_lines)
    out_lines.extend(SUFFIX.splitlines())
    write_lines(target, out_lines)

def main():
    try:
        directory, filename_no_ext, source_file = ask_inputs()
        unique_ids, _ = unique_trace_ids_with_progress_xlsx(source_file)
        Count = len(unique_ids)
        print(f"Всего уникальных объектов: {Count}")

        # Параметры группировки
        objects_per_group = 3500
        groupsCount = math.ceil(Count / objects_per_group) if Count > 0 else 0

        output_dir = directory / "output"
        output_dir.mkdir(parents=True, exist_ok=True)

        # Формирование групп напрямую (без промежуточных файлов)
        current_group = 1
        objects_in_group = 0
        buffer_lines: List[str] = []

        for i, tid in enumerate(unique_ids, start=1):
            buffer_lines.extend(object_dsl_lines_for(tid))
            objects_in_group += 1

            end_of_data = (i == Count)
            if objects_in_group >= objects_per_group or end_of_data:
                target = output_dir / f"dsl_group-{current_group}.txt"
                write_group_file_wrapped(target, buffer_lines)
                buffer_lines = []
                objects_in_group = 0
                current_group += 1
                progress_bar(min(current_group - 1, groupsCount), groupsCount, prefix="Обёртка файлов: ")
                time.sleep(0.005)

        progress_bar(groupsCount, groupsCount, prefix="Обёртка файлов: ")
        sys.stdout.write("\n")

        print(f"Преобразование файлов завершено. Всего - {groupsCount} файлов.")
    except KeyboardInterrupt:
        print("\nОстановлено пользователем.")

if __name__ == "__main__":
    main()
