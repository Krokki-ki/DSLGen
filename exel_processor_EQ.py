#!/usr/bin/env python3
from __future__ import annotations

import sys
from pathlib import Path
from typing import Iterable, List, Tuple

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("Для работы требуется пакет 'openpyxl'. Установите: pip install openpyxl")
    sys.exit(1)

# ---------------------------
# Вспомогательные функции
# ---------------------------
def progress_bar(current: int, total: int, prefix: str = "", width: int = 30) -> None:
    if total <= 0:
        pct = 0.0
    else:
        pct = min(100.0, (current / total) * 100.0)
    filled = int(width * pct / 100)
    bar = "#" * filled + "." * (width - filled)
    sys.stdout.write(f"\r{prefix}[{bar}] {pct:5.1f}%")
    sys.stdout.flush()

def ask_directory() -> Path:
    directory_str = input("Укажите путь директории к папке, в которой находятся искомые XLSX файлы: ").strip()
    directory = Path(directory_str).expanduser().resolve()
    if not directory.is_dir():
        print("Указана неверная директория! Проверьте расположение файлов XLSX и повторите попытку!")
        return ask_directory()
    return directory

def validate_directory_xlsx_only(directory: Path, produced_files: set[str]) -> Tuple[List[Path], int]:
    """
    Разрешены только .xlsx и создаваемые этим модулем файлы (из produced_files).
    Возвращает (список XLSX к обработке, count_files).
    """
    files = [p for p in directory.iterdir() if p.is_file()]
    xlsx_files: List[Path] = []
    for f in files:
        if f.name in produced_files:
            continue
        if f.suffix.lower() != ".xlsx":
            print("Указана неверная директория! Проверьте расположение файлов XLSX и повторите попытку!")
            raise RuntimeError("BAD_DIR")
        xlsx_files.append(f)
    return sorted(xlsx_files), len(xlsx_files)

def create_excel_with_headers(path: Path) -> None:
    """
    Порядок колонок:
    F0UCUS1, F0UIDPL, F0UEAN, F0USCRD, F0USCON, SCACT
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "EQ_data"
    ws.append(["F0UCUS1", "F0UIDPL", "F0UEAN", "F0USCRD", "F0USCON", "SCACT"])
    wb.save(path)

def append_rows_to_excel(xlsx_path: Path, rows: Iterable[List[str]]) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(xlsx_path)

# ---------------------------
# Чтение XLSX и сбор всех строк (устойчивое к «шапкам» и формату)
# ---------------------------
EXPECTED_HEADERS = ["F0UCUS1", "F0UIDPL", "F0UEAN", "F0USCRD", "F0USCON", "SCACT"]

def load_ws_robust(path: Path):
    # Пытаемся читать в read_only, если лист «пустой» — читаем полноценно
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    try:
        dim = ws.calculate_dimension()
        if dim == "A1:A1":  # признак «не определённой» области
            raise ValueError("unsized")
    except Exception:
        wb = load_workbook(path, data_only=True, read_only=False)
        ws = wb.active
    return ws

def normalize_header_cell(x: object) -> str:
    return ("" if x is None else str(x)).strip().upper()

def read_rows_from_xlsx(path: Path) -> List[List[str]]:
    """
    Читает все строки данных из XLSX:
    - пропускает ведущие пустые строки;
    - находит первую непустую строку и считает её шапкой;
    - сопоставляет колонки по именам EXPECTED_HEADERS (case-insensitive, trim);
    - возвращает значения только по этим шести полям в заданном порядке.
    """
    ws = load_ws_robust(path)

    # Найти строку-шапку: первую строку, где есть хотя бы одно непустое значение
    header_row_idx = None
    raw_header: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [("" if v is None else str(v).strip()) for v in row]
        if any(values):
            raw_header = values
            header_row_idx = i
            break
    if header_row_idx is None:
        return []

    # Построить маппинг имя столбца (нормализованное) -> индекс
    norm_header = [normalize_header_cell(h) for h in raw_header]
    name_to_idx = {name: idx for idx, name in enumerate(norm_header) if name}

    # Сопоставить требуемые колонки
    wanted_norm = [h.upper() for h in EXPECTED_HEADERS]
    missing = [h for h in wanted_norm if h not in name_to_idx]
    if missing:
        print(f"Предупреждение: {path.name}: отсутствуют колонки {missing}. Будут заполнены пустыми значениями.")

    # Читать данные со следующей строки после шапки
    rows_out: List[List[str]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= header_row_idx:
            continue
        values = [("" if v is None else str(v).strip()) for v in row]

        picked: list[str] = []
        for want in wanted_norm:
            if want in name_to_idx:
                idx = name_to_idx[want]
                picked.append(values[idx] if idx < len(values) else "")
            else:
                picked.append("")

        # Пропуск полностью пустых строк
        if all(v == "" for v in picked):
            continue

        # Гарантируем ровно 6 полей
        if len(picked) > 6:
            picked = picked[:6]
        elif len(picked) < 6:
            picked += [""] * (6 - len(picked))

        rows_out.append(picked)

    return rows_out

# ---------------------------
# Дедупликация строк по всем колонкам (с сохранением порядка)
# ---------------------------
def deduplicate_rows(rows: List[List[str]]) -> List[List[str]]:
    """
    Удаляет полные дубликаты строк (по всем 6 полям), оставляя первое вхождение.
    Порядок исходных уникальных строк сохраняется.
    """
    seen = set()
    out: List[List[str]] = []
    for r in rows:
        key = tuple(r)  # кортеж 6 полей
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out

# ---------------------------
# Основная логика
# ---------------------------
def main():
    while True:
        try:
            directory = ask_directory()
            xlsx_name = "exel_customer_data_EQ.xlsx"
            xlsx_path = directory / xlsx_name
            produced_files = {xlsx_name}

            try:
                xlsx_files, count_files = validate_directory_xlsx_only(directory, produced_files)
            except RuntimeError as e:
                if str(e) == "BAD_DIR":
                    continue
                else:
                    raise

            if count_files == 0:
                print("XLSX файлы не найдены в указанной директории. Повторите попытку.")
                continue

            # Читаем все файлы и собираем все строки
            print("Чтение XLSX-файлов и сбор строк...")
            all_rows: List[List[str]] = []
            total = count_files
            processed = 0

            for p in xlsx_files:
                processed += 1
                progress_bar(processed, total, prefix="Прогресс: ")
                rows = read_rows_from_xlsx(p)
                if rows:
                    all_rows.extend(rows)

            progress_bar(total, total, prefix="Прогресс: ")
            sys.stdout.write("\n")

            print(f"Считано строк всего: {len(all_rows)} до дедупликации")

            # Дедупликация по всем полям
            print("Дедупликация строк по всем полям (оставляются уникальные строки)...")
            unique_rows = deduplicate_rows(all_rows)
            removed = len(all_rows) - len(unique_rows)
            print(f"Удалено дубликатов: {removed}. Уникальных строк: {len(unique_rows)}")

            # Создаем итоговый Excel и записываем уникальные строки
            if xlsx_path.exists():
                xlsx_path.unlink()
            create_excel_with_headers(xlsx_path)

            print("Запись итогового EXEL-файла...")
            append_rows_to_excel(xlsx_path, unique_rows)
            print(f'Готово. Итоговый EXEL-файл: "{xlsx_name}" в директории {directory}')

            break

        except KeyboardInterrupt:
            print("\nОстановлено пользователем.")
            break
        except RuntimeError as e:
            if str(e) == "OPENPYXL_MISSING":
                break
            raise

if __name__ == "__main__":
    main()
