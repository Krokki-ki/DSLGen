#!/usr/bin/env python3
from __future__ import annotations

import sys
import warnings
from pathlib import Path
from typing import Dict, List, Tuple, Optional

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    print("Для работы summator.py требуется пакет 'openpyxl'. Установите: pip install openpyxl")
    sys.exit(1)

def progress_bar(current: int, total: int, prefix: str = "", width: int = 30) -> None:
    if total <= 0:
        pct = 0.0
    else:
        pct = min(100.0, (current / total) * 100.0)
    filled = int(width * pct / 100)
    bar = "#" * filled + "." * (width - filled)
    sys.stdout.write(f"\r{prefix}[{bar}] {pct:5.1f}%")
    sys.stdout.flush()

DISCLAIMER = (
    "Убедитесь, что обе EXEL-таблицы (из шага 2 и из шага 4), необходимые для соединения в единый файл, находятся в одной директории (желательно в отдельной папке)\n"
    "Должно быть только два файла.\n"
    "\n"
    "Чтобы продолжить - введите \"ДА\"\n"
    "Чтобы вернуться в главное меню - введите 0"
)

# Ожидаемые заголовки
KIBANA_SET = {"customerId", "cardId", "account", "terminalId"}
EQ_BASE_SET = {"F0UCUS1", "F0UIDPL", "F0UEAN", "F0USCRD", "F0USCON"}
EQ_FULL_SET = {"F0UCUS1", "F0UIDPL", "F0UEAN", "F0USCRD", "F0USCON", "SCACT"}

def read_headers_xlsx(path: Path) -> List[str]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=UserWarning)
        wb = load_workbook(path)
    ws = wb.active
    headers: List[str] = []
    for cell in ws[1]:
        headers.append(str(cell.value) if cell.value is not None else "")
    return headers

def classify_file(path: Path) -> Optional[str]:
    """
    Возвращает 'kibana' или 'eq' по заголовкам, иначе None.
    """
    headers = [h.strip() for h in read_headers_xlsx(path)]
    set_h = set(headers)
    if KIBANA_SET.issubset(set_h):
        return "kibana"
    # EQ: допускаем варианты с/без SCACT
    if EQ_BASE_SET.issubset(set_h) or EQ_FULL_SET.issubset(set_h):
        return "eq"
    return None

def ask_disclaimer_and_directory() -> Optional[Path]:
    print(DISCLAIMER)
    ans = input("> ").strip()
    if ans == "0":
        return None
    if ans.lower() != "да":
        print("Ввод не распознан. Возврат в главное меню.")
        return None
    directory_str = input("Укажите общую директорию, в которой содержатся два EXEL-файла: ").strip()
    directory = Path(directory_str).expanduser().resolve()
    if not directory.is_dir():
        print("Указанная директория недоступна. Возврат в главное меню.")
        return None
    return directory

def list_two_xlsx(directory: Path) -> List[Path]:
    files = [p for p in directory.iterdir() if p.is_file() and p.suffix.lower() == ".xlsx"]
    return sorted(files)

def rename_columns_kibana(headers: List[str]) -> Dict[str, int]:
    """
    Kibana (новая структура):
      customerId -> kibana_cus
      cardId     -> kibana_card
      account    -> kibana_account
      terminalId -> kibana_terminal
    """
    idx: Dict[str, int] = {}
    for i, h in enumerate(headers):
        hn = (h or "").strip()
        if hn == "customerId":
            idx["kibana_cus"] = i
        elif hn == "cardId":
            idx["kibana_card"] = i
        elif hn == "account":
            idx["kibana_account"] = i
        elif hn == "terminalId":
            idx["kibana_terminal"] = i
    return idx

def rename_columns_eq(headers: List[str]) -> Dict[str, int]:
    """
    EQ:
      F0UCUS1 -> eq_cus
      F0UIDPL -> eq_card
      F0UEAN  -> eq_account
      F0USCRD -> eq_type
      F0USCON -> eq_contract
      SCACT   -> eq_scact (если есть)
    """
    idx: Dict[str, int] = {}
    for i, h in enumerate(headers):
        hn = (h or "").strip()
        if hn == "F0UCUS1":
            idx["eq_cus"] = i
        elif hn == "F0UIDPL":
            idx["eq_card"] = i
        elif hn == "F0UEAN":
            idx["eq_account"] = i
        elif hn == "F0USCRD":
            idx["eq_type"] = i
        elif hn == "F0USCON":
            idx["eq_contract"] = i
        elif hn == "SCACT":
            idx["eq_scact"] = i
    return idx

def load_rows(path: Path) -> Tuple[List[str], List[List[object]]]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=UserWarning)
        wb = load_workbook(path)
    ws = wb.active
    rows: List[List[object]] = []
    headers: List[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(v) if v is not None else "" for v in row]
        else:
            rows.append(list(row))
    return headers, rows

def build_map(rows: List[List[object]], col_index: int, select_cols: List[Optional[int]]) -> Dict[object, List[object]]:
    m: Dict[object, List[object]] = {}
    for r in rows:
        if col_index < 0 or col_index >= len(r):
            continue
        key = r[col_index]
        values: List[object] = []
        for ci in select_cols:
            v = r[ci] if ci is not None and 0 <= ci < len(r) else None
            values.append(v)
        m[key] = values
    return m

def full_outer_join(left_map: Dict[object, List[object]], right_map: Dict[object, List[object]], left_width: int, right_width: int) -> List[List[object]]:
    result: List[List[object]] = []
    keys = set(left_map.keys()) | set(right_map.keys())
    for k in keys:
        left_vals = left_map.get(k)
        right_vals = right_map.get(k)
        if left_vals is None:
            left_vals = [None] * left_width
        if right_vals is None:
            right_vals = [None] * right_width
        result.append(left_vals + right_vals)
    return result

def write_output(path: Path, headers: List[str], rows: List[List[object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "join"
    ws.append(headers)
    total = len(rows)
    for i, r in enumerate(rows, start=1):
        ws.append(r)
        if i % 100 == 0 or i == total:
            progress_bar(i, total, prefix="Сохранение результата: ")
    if total == 0:
        progress_bar(1, 1, prefix="Сохранение результата: ")
    sys.stdout.write("\n")
    wb.save(path)

def ask_key() -> Tuple[str, str, str]:
    """
    Возвращает кортеж из:
      (key_name_for_output, kibana_col_key, eq_col_key)
    """
    print(
        'Укажите название ключа, по которому хотите выполнить объединение двух таблиц (выберите один из режимов ниже и введите):\n'
        '"1" - объединение по CUS\n'
        '"2" - объединение по Card\n'
        '"3" - объединение по Account\n'
    )
    choice = input("> ").strip()
    if choice == "1":
        return ("CUS", "kibana_cus", "eq_cus")
    elif choice == "2":
        return ("Card", "kibana_card", "eq_card")
    elif choice == "3":
        return ("Account", "kibana_account", "eq_account")
    else:
        print("Ввод не распознан. Возврат в главное меню.")
        return ("", "", "")

def is_nonempty(v) -> bool:
    return v is not None and str(v).strip() != ""

def sort_joined_file(xlsx_path: Path, eq_width: int, kib_width: int) -> None:
    """
    Сортирует по группам заполненности:
      0 — EQ и Kibana блоки полностью заполнены
      1 — EQ полностью, Kibana пуст
      2 — Kibana полностью, EQ пуст
      3 — иные случаи
    eq_width: ширина EQ-блока (5 или 6, если есть eq_scact)
    kib_width: ширина Kibana-блока (всегда 4)
    """
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=UserWarning)
        wb = load_workbook(xlsx_path)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.save(xlsx_path)
        return

    headers = list(rows[0])
    data = [list(r) for r in rows[1:]]

    def group_row(row: List[object]) -> int:
        eq_vals = row[0:eq_width]
        kb_vals = row[eq_width:eq_width + kib_width]
        eq_all = all(is_nonempty(x) for x in eq_vals)
        kb_all = all(is_nonempty(x) for x in kb_vals)
        eq_any = any(is_nonempty(x) for x in eq_vals)
        kb_any = any(is_nonempty(x) for x in kb_vals)
        if eq_all and kb_all:
            return 0
        if eq_all and not kb_any:
            return 1
        if kb_all and not eq_any:
            return 2
        return 3

    grouped = {0: [], 1: [], 2: [], 3: []}
    for r in data:
        grouped[group_row(r)].append(r)
    sorted_data = grouped[0] + grouped[1] + grouped[2] + grouped[3]

    ws.delete_rows(1, ws.max_row)
    ws.append(headers)

    total = len(sorted_data)
    for i, r in enumerate(sorted_data, start=1):
        ws.append(r)
        if i % 200 == 0 or i == total:
            progress_bar(i, total, prefix="Сортировка результата: ")
    if total == 0:
        progress_bar(1, 1, prefix="Сортировка результата: ")
    sys.stdout.write("\n")
    wb.save(xlsx_path)

# ---------- Основная логика ----------
def main():
    # Дисклеймер и директория
    directory = ask_disclaimer_and_directory()
    if directory is None:
        return

    # Ожидаем ровно 2 XLSX файла
    xlsx_files = list_two_xlsx(directory)
    if len(xlsx_files) != 2:
        print("В директории должно быть ровно два XLSX-файла. Возврат в главное меню.")
        return

    # Определяем какой из двух файлов есть kibana/eq (с учетом новой структуры Kibana и опц. SCACT на EQ)
    roles: Dict[str, Path] = {}
    for p in xlsx_files:
        label = classify_file(p)
        if label is None:
            print(f"Не удалось распознать структуру файла: {p.name}. Проверьте заголовки.")
            return
        roles[label] = p

    if "kibana" not in roles or "eq" not in roles:
        print("Не удалось однозначно определить 'kibana' и 'eq' файлы по заголовкам. Проверьте входные файлы.")
        return

    kibana_path = roles["kibana"]
    eq_path = roles["eq"]

    # Выбор ключа
    key_name, kib_key, eq_key = ask_key()
    if not key_name:
        return

    print("Выполняется обработка EXEL файлов")
    progress_bar(0, 2, prefix="Прогресс: ")

    # Загружаем данные из обоих файлов
    kib_headers_raw, kib_rows_raw = load_rows(kibana_path)
    progress_bar(1, 2, prefix="Прогресс: ")
    eq_headers_raw, eq_rows_raw = load_rows(eq_path)
    progress_bar(2, 2, prefix="Прогресс: ")
    sys.stdout.write("\n")

    # Построение индексов колонок
    kib_idx = rename_columns_kibana(kib_headers_raw)
    eq_idx = rename_columns_eq(eq_headers_raw)

    # Проверка наличия обязательных колонок
    required_kib = {"kibana_cus", "kibana_card", "kibana_account", "kibana_terminal"}
    required_eq = {"eq_cus", "eq_card", "eq_account", "eq_type", "eq_contract"}  # eq_scact — опционально
    if not required_kib.issubset(set(kib_idx.keys())):
        print("В файле Kibana отсутствуют необходимые колонки (ожидались customerId, cardId, account, terminalId).")
        return
    if not required_eq.issubset(set(eq_idx.keys())):
        print("В файле EQ отсутствуют необходимые колонки (ожидались F0UCUS1, F0UIDPL, F0UEAN, F0USCRD, F0USCON).")
        return

    kib_key_index = kib_idx.get(kib_key, -1)
    eq_key_index = eq_idx.get(eq_key, -1)
    if kib_key_index < 0 or eq_key_index < 0:
        print("Не удалось определить ключевые колонки для объединения. Проверьте входные файлы и выбранный режим.")
        return

    # Выбор колонок для результата
    eq_select: List[int] = [
        eq_idx["eq_cus"],
        eq_idx["eq_card"],
        eq_idx["eq_account"],
        eq_idx["eq_type"],
        eq_idx["eq_contract"],
    ]
    if "eq_scact" in eq_idx:
        eq_select.append(eq_idx["eq_scact"])

    kib_select: List[int] = [
        kib_idx["kibana_cus"],
        kib_idx["kibana_card"],
        kib_idx["kibana_account"],
        kib_idx["kibana_terminal"],
    ]

    # Построение map по ключу
    eq_map = build_map(eq_rows_raw, eq_key_index, eq_select)
    kib_map = build_map(kib_rows_raw, kib_key_index, kib_select)

    # Объединение (полное внешнее)
    print(f'Выполняется объединение таблиц по ключу {key_name}...')
    joined_rows = full_outer_join(eq_map, kib_map, left_width=len(eq_select), right_width=len(kib_select))

    # Заголовки результата: EQ-блок (5 или 6) + Kibana-блок (4)
    out_headers = ["eq_cus", "eq_card", "eq_account", "eq_type", "eq_contract"]
    if len(eq_select) == 6:
        out_headers.append("eq_scact")
    out_headers += ["kibana_cus", "kibana_card", "kibana_account", "kibana_terminal"]

    # Сохранение результата
    out_name = f"All_join_data-{key_name}.xlsx"
    out_path = directory / out_name
    write_output(out_path, out_headers, joined_rows)

    # Сортировка с учетом реальной ширины блоков
    sort_joined_file(out_path, eq_width=len(eq_select), kib_width=len(kib_select))

    print(f'Объединение таблиц завершено. Файл "{out_name}" выгружен в директорию {directory}')

if __name__ == "__main__":
    main()
