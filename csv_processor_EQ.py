#!/usr/bin/env python3
from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Iterable, List, Tuple, Optional

# Excel
try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None

# ---------------------------
# Вспомогательные функции
# ---------------------------
def progress_bar(current: int, total: int, prefix: str = "", width: int = 30) -> None:
    if total <= 0:
        pct = 0.0
    else:
        pct = min(100.0, (current / total) * 100.0)
    filled = int(width * pct / 100)
    bar = "#" * filled + "." * (width - filled)
    sys.stdout.write(f"\r{prefix}[{bar}] {pct:5.1f}%")
    sys.stdout.flush()

def ask_directory() -> Path:
    directory_str = input("Укажите путь директории к папке, в которой находятся искомые CSV файлы: ").strip()
    directory = Path(directory_str).expanduser().resolve()
    if not directory.is_dir():
        print("Указана неверная директория! Проверьте расположение файлов CSV и повторите попытку!")
        return ask_directory()
    return directory

def validate_directory_csv_only(directory: Path, produced_files: set[str]) -> Tuple[List[Path], int]:
    """
    Проверяет, что в директории нет инородных файлов (разрешены .csv и создаваемые модулем файлы).
    Возвращает список CSV и их число.
    """
    files = [p for p in directory.iterdir() if p.is_file()]
    csv_files: List[Path] = []
    for f in files:
        if f.name in produced_files:
            continue
        if f.suffix.lower() != ".csv":
            print("Указана неверная директория! Проверьте расположение файлов CSV и повторите попытку!")
            raise RuntimeError("BAD_DIR")
        csv_files.append(f)
    return sorted(csv_files), len(csv_files)

def ensure_openpyxl_available() -> None:
    if Workbook is None or load_workbook is None:
        print("Для формирования Excel-файла требуется пакет 'openpyxl'. Установите его: pip install openpyxl")
        raise RuntimeError("OPENPYXL_MISSING")

def create_excel_with_headers(path: Path) -> None:
    """
    Порядок колонок:
    F0UCUS1, F0UIDPL, F0UEAN, F0USCRD, F0USCON, SCACT
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "EQ_data"
    ws.append(["F0UCUS1", "F0UIDPL", "F0UEAN", "F0USCRD", "F0USCON", "SCACT"])
    wb.save(path)

def append_rows_to_excel(xlsx_path: Path, rows: Iterable[Tuple[Optional[str], Optional[str], Optional[str], Optional[str], Optional[str], Optional[str]]]) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb.active
    for row in rows:
        ws.append(list(row))
    wb.save(xlsx_path)

# ---------------------------
# Регулярные выражения для извлечения
# ---------------------------
# a) F0UCUS1: [A-Z0-9]{6}
RE_F0UCUS1 = re.compile(r"\b([A-Z0-9]{6})\b")

# b) F0UIDPL: [0-9]{8,11}
RE_F0UIDPL = re.compile(r"\b([0-9]{8,11})\b")

# c) F0UEAN: [0-9]{18,22}
RE_F0UEAN = re.compile(r"\b([0-9]{18,22})\b")

# d) F0USCRD: [A-Z0-9]{2,3}
RE_F0USCRD = re.compile(r"\b([A-Z0-9]{2,3})\b")

# e) F0USCON: [A-Z0-9]{4,7}
RE_F0USCON = re.compile(r"\b([A-Z0-9]{4,7})\b")

# f) SCACT: [A-Z]{2,3}
RE_SCACT = re.compile(r"\b([A-Z]{2,3})\b")

def extract_first_or_none(pattern: re.Pattern, text: str) -> Optional[str]:
    m = pattern.search(text)
    return m.group(1) if m else None

def extract_row_from_text(text: str) -> Tuple[Optional[str], Optional[str], Optional[str], Optional[str], Optional[str], Optional[str]]:
    """
    Для заданной текстовой строки извлекает первое вхождение каждого поля по его regex.
    Возвращает кортеж значений для колонок в порядке Excel:
    (F0UCUS1, F0UIDPL, F0UEAN, F0USCRD, F0USCON, SCACT)
    Если в строке какие-то поля не найдены — возвращает None в соответствующей позиции.
    """
    f0ucus1 = extract_first_or_none(RE_F0UCUS1, text)
    f0uidpl = extract_first_or_none(RE_F0UIDPL, text)
    f0uean  = extract_first_or_none(RE_F0UEAN,  text)
    f0uscrd = extract_first_or_none(RE_F0USCRD, text)
    f0uscon = extract_first_or_none(RE_F0USCON, text)
    scact   = extract_first_or_none(RE_SCACT,   text)
    return (f0ucus1, f0uidpl, f0uean, f0uscrd, f0uscon, scact)

# ---------------------------
# Основной конвейер
# ---------------------------
def main():
    while True:
        try:
            directory = ask_directory()
            xlsx_name = "exel_customer_data_EQ.xlsx"
            xlsx_path = directory / xlsx_name

            produced_files = {xlsx_name}

            try:
                csv_files, count_files = validate_directory_csv_only(directory, produced_files)
            except RuntimeError as e:
                if str(e) == "BAD_DIR":
                    continue
                else:
                    raise

            if count_files == 0:
                print("CSV файлы не найдены в указанной директории. Повторите попытку.")
                continue

            ensure_openpyxl_available()
            if xlsx_path.exists():
                xlsx_path.unlink()
            create_excel_with_headers(xlsx_path)

            print("Обработка файлов для генерации EXEL")
            total = count_files
            processed = 0

            for csv_path in csv_files:
                processed += 1
                progress_bar(processed, total, prefix="Прогресс: ")

                rows_batch: List[Tuple[Optional[str], Optional[str], Optional[str], Optional[str], Optional[str], Optional[str]]] = []
                with csv_path.open("r", encoding="utf-8", errors="ignore") as f:
                    for line in f:
                        # Из каждой строки CSV извлекаем первое вхождение каждого шаблона
                        row = extract_row_from_text(line)
                        # Если нужны только полные строки, можно включить фильтр ниже:
                        # if all(v is not None for v in row):
                        #     rows_batch.append(row)
                        # else:
                        #     continue
                        rows_batch.append(row)

                if rows_batch:
                    append_rows_to_excel(xlsx_path, rows_batch)

            progress_bar(total, total, prefix="Прогресс: ")
            sys.stdout.write("\n")
            print(f'Готово. Итоговый EXEL-файл: "{xlsx_name}"')

            break

        except KeyboardInterrupt:
            print("\nОстановлено пользователем.")
            break
        except RuntimeError as e:
            if str(e) == "OPENPYXL_MISSING":
                break
            raise

if __name__ == "__main__":
    main()
