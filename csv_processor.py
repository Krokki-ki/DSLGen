#!/usr/bin/env python3
from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Iterable, Iterator, Tuple, List

# Для Excel
try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None

# ---------------------------
# Вспомогательные функции
# ---------------------------
def progress_bar(current: int, total: int, prefix: str = "", width: int = 30) -> None:
    if total <= 0:
        pct = 0.0
    else:
        pct = min(100.0, (current / total) * 100.0)
    filled = int(width * pct / 100)
    bar = "#" * filled + "." * (width - filled)
    sys.stdout.write(f"\r{prefix}[{bar}] {pct:5.1f}%")
    sys.stdout.flush()

def ask_directory() -> Path:
    directory_str = input("Укажите путь директории к папке, в которой находятся искомые CSV файлы: ").strip()
    directory = Path(directory_str).expanduser().resolve()
    if not directory.is_dir():
        print("Указана неверная директория! Проверьте расположение файлов CSV и повторите попытку!")
        return ask_directory()
    return directory

def validate_directory_has_only_csv(directory: Path, txt_name: str, xlsx_name: str) -> Tuple[list[Path], int]:
    """
    Разрешены только .csv, .txt (для итогового txt) и .xlsx (для итоговой таблицы).
    Любой иной файл => ошибка и возврат к началу.
    Возвращает (список CSV к обработке, count_files).
    """
    allowed = {".csv", ".txt", ".xlsx"}
    files = [p for p in directory.iterdir() if p.is_file()]
    csv_files: list[Path] = []
    for f in files:
        ext = f.suffix.lower()
        if ext not in allowed:
            print("Указана неверная директория! Проверьте расположение файлов CSV и повторите попытку!")
            raise RuntimeError("BAD_DIR")
        # Исключить итоговые файлы из обработки
        if f.name in {txt_name, xlsx_name}:
            continue
        if ext == ".csv":
            csv_files.append(f)
    return sorted(csv_files), len(csv_files)

def write_lines(path: Path, lines: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for s in lines:
            f.write(s)
            if not s.endswith("\n"):
                f.write("\n")

def append_lines(path: Path, lines: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as f:
        for s in lines:
            f.write(s)
            if not s.endswith("\n"):
                f.write("\n")

def deduplicate_file_in_place(path: Path) -> tuple[int, int]:
    # Возвращает (count_duplicates, count_unique)
    seen: set[str] = set()
    uniques: list[str] = []
    duplicates = 0
    with path.open("r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            s = line.strip()
            if not s:
                continue
            if s in seen:
                duplicates += 1
            else:
                seen.add(s)
                uniques.append(s)
    write_lines(path, uniques)
    return duplicates, len(uniques)

def strip_prefix_in_place(path: Path, filter_key: str) -> None:
    # Удаляет префикс 'filter=' у каждой строки
    out: list[str] = []
    prefix = f"{filter_key}="
    with path.open("r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            s = line.strip()
            if not s:
                continue
            if s.startswith(prefix):
                out.append(s[len(prefix):])
            else:
                out.append(s)
    write_lines(path, out)

def extract_matches_for_txt(line: str, filter_key: str, capture_regex: re.Pattern) -> Iterator[str]:
    # Ищем все вхождения вида filter=<VALUE> по шаблону VALUE
    for m in re.finditer(re.escape(filter_key) + r"=(" + capture_regex.pattern + r")", line):
        value = m.group(1)
        yield f"{filter_key}={value}"

# ---------------------------
# TXT-конвейер (csv_general.txt)
# ---------------------------
def run_txt_pipeline(directory_csv: Path, filter_key: str, txt_path: Path, csv_files: list[Path]) -> Tuple[int, int, int]:
    """
    Обработка TXT-конвейера:
    - собираем filter=VALUE из всех CSV
    - пишем в csv_general.txt
    - дедупликация и удаление префикса
    Возвращает (count_dubles, filter_count_unique, count_files).
    """
    # Создаём/очищаем файл
    if txt_path.exists():
        txt_path.unlink()
    txt_path.touch()

    print("Обработка файлов для генерации txt")
    value_regex = re.compile(r"[A-Z0-9]+")
    count_files = len(csv_files)
    processed = 0

    for csv_path in csv_files:
        processed += 1
        progress_bar(processed, count_files, prefix="Прогресс: ")
        # Парсим построчно как обычный текст
        with csv_path.open("r", encoding="utf-8", errors="ignore") as f:
            buffer_out: list[str] = []
            for line in f:
                for match_line in extract_matches_for_txt(line, filter_key, value_regex):
                    buffer_out.append(match_line)
            if buffer_out:
                append_lines(txt_path, buffer_out)

    progress_bar(count_files, count_files, prefix="Прогресс: ")
    sys.stdout.write("\n")

    print("Обработка CSV-файлов завершена. Начинается процесс дедубликации и форматирования:")
    count_dubles, filter_count_unique = deduplicate_file_in_place(txt_path)
    strip_prefix_in_place(txt_path, filter_key)

    print(f'Форматирование завершено. Найдено {count_dubles} дублей.')
    print(f'Уникальных значений параметра: {filter_key} {filter_count_unique}')
    print(f'Файлов обработано: {count_files}')

    return count_dubles, filter_count_unique, count_files

# ---------------------------
# Excel-конвейер (exel_customer_data.xlsx)
# ---------------------------
def ensure_openpyxl_available() -> None:
    if Workbook is None or load_workbook is None:
        print("Для формирования Excel-файла требуется пакет 'openpyxl'. Установите его: pip install openpyxl")
        raise RuntimeError("OPENPYXL_MISSING")

def create_empty_excel(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    # Новый порядок: customerId, cardId, account, terminalId
    ws.append(["customerId", "cardId", "account", "terminalId"])
    wb.save(path)

def extract_four_filters_from_line(line: str) -> list[tuple[str, str]]:
    """
    Новые правила извлечения (порядок: customerId, cardId, account, terminalId):

      a) customerId=[A-Z0-9]{6}
         пример: customerId=A9WH99

      b) cardId=[0-9]+
         пример: cardId=46848486841

      c) account=[0-9]+
         пример: account=98178102068106464682

      d) terminalId=[0-9]+
         пример: terminalId=228038

    Возвращает список пар (label, raw_label_value), если все 4 параметра найдены; иначе возвращает [].
    """
    results: list[tuple[str, str]] = []

    # a) customerId
    m1 = re.search(r"(customerId=)([A-Z0-9]{6})\b", line)
    if not m1:
        return []
    results.append(("customerId", m1.group(0)))  # 'customerId=VALUE'

    # b) cardId
    m2 = re.search(r"(cardId=)([0-9]+)\b", line)
    if not m2:
        return []
    results.append(("cardId", m2.group(0)))

    # c) account
    m3 = re.search(r"(account=)([0-9]+)\b", line)
    if not m3:
        return []
    results.append(("account", m3.group(0)))

    # d) terminalId
    m4 = re.search(r"(terminalId=)([0-9]+)\b", line)
    if not m4:
        return []
    results.append(("terminalId", m4.group(0)))

    return results

def append_rows_to_excel(xlsx_path: Path, rows: Iterable[tuple[str, str, str, str]]) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb.active
    for row in rows:
        ws.append(list(row))
    wb.save(xlsx_path)

def strip_prefixes_in_excel(xlsx_path: Path) -> None:
    """
    Удаляем префиксы 'customerId=', 'cardId=', 'account=', 'terminalId='
    из ячеек соответствующих столбцов (в новом порядке колонок).
    """
    wb = load_workbook(xlsx_path)
    ws = wb.active

    def strip_prefix(value: str, prefix: str) -> str:
        if isinstance(value, str) and value.startswith(prefix):
            return value[len(prefix):]
        return value

    for _idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        # Порядок: A=customerId, B=cardId, C=account, D=terminalId
        if row[0].value:
            row[0].value = strip_prefix(row[0].value, "customerId=")
        if row[1].value:
            row[1].value = strip_prefix(row[1].value, "cardId=")
        if row[2].value:
            row[2].value = strip_prefix(row[2].value, "account=")
        if row[3].value:
            row[3].value = strip_prefix(row[3].value, "terminalId=")

    wb.save(xlsx_path)

def excel_remove_duplicates_inplace(xlsx_path: Path) -> int:
    """
    Удаляет полностью дублирующиеся строки (по всем столбцам) в файле xlsx.
    Возвращает количество удалённых дублей.
    """
    wb = load_workbook(xlsx_path)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.save(xlsx_path)
        return 0

    headers = rows[0]
    data = rows[1:]

    seen = set()
    unique_rows = []
    removed = 0

    for r in data:
        key = tuple("" if v is None else str(v) for v in r)
        if key in seen:
            removed += 1
        else:
            seen.add(key)
            unique_rows.append(list(r))

    ws.delete_rows(1, ws.max_row)
    ws.append(headers)

    total = len(unique_rows)
    for i, r in enumerate(unique_rows, start=1):
        ws.append(r)
        if i % 200 == 0 or i == total:
            progress_bar(i, total, prefix="Дедубликация EXEL: ")
    if total == 0:
        progress_bar(1, 1, prefix="Дедубликация EXEL: ")
    sys.stdout.write("\n")

    wb.save(xlsx_path)
    return removed

def run_excel_dedup_with_notice(xlsx_path: Path) -> None:
    print("Запущена дедубликация строк EXEL-таблицы (exel_customer_data.xlsx)...")
    removed = excel_remove_duplicates_inplace(xlsx_path)
    print(f"Дедубликация завершена. Удалено дублирующихся строк: {removed}")

# ---------------------------
# Основная логика
# ---------------------------
def main():
    while True:
        try:
            directory_csv = ask_directory()

            # Подготовка путей итоговых файлов
            txt_path = directory_csv / "csv_general.txt"
            xlsx_path = directory_csv / "exel_customer_data.xlsx"

            # Валидация каталога и сбор CSV к обработке
            try:
                csv_files, count_files = validate_directory_has_only_csv(directory_csv, txt_path.name, xlsx_path.name)
            except RuntimeError as e:
                if str(e) == "BAD_DIR":
                    continue
                else:
                    raise

            if count_files == 0:
                print("CSV файлы не найдены в указанной директории. Повторите попытку.")
                continue

            # TXT-конвейер: подсказка фильтра без изменений
            filter_key = input(
                "Укажите параметр, по которму требуется отобрать данные во всех файлах CSV (например, customerId, cardId, account): "
            ).strip()
            if not filter_key:
                print("Параметр фильтра не может быть пустым. Повторите ввод.")
                continue

            _count_dubles, _filter_count_unique, _ = run_txt_pipeline(directory_csv, filter_key, txt_path, csv_files)

            # EXCEL-конвейер: создаём и заполняем по новым правилам (customerId, cardId, account, terminalId)
            ensure_openpyxl_available()
            if xlsx_path.exists():
                xlsx_path.unlink()
            create_empty_excel(xlsx_path)

            print("Обработка файлов дял генерации exel")
            total = len(csv_files)
            processed = 0
            for csv_path in csv_files:
                processed += 1
                progress_bar(processed, total, prefix="Прогресс: ")
                rows_to_add: list[tuple[str, str, str, str]] = []
                with csv_path.open("r", encoding="utf-8", errors="ignore") as f:
                    for line in f:
                        pairs = extract_four_filters_from_line(line)
                        if not pairs:
                            continue
                        mapping = {k: v for k, v in pairs}
                        # новые 4 поля в требуемом порядке
                        if all(k in mapping for k in ("customerId", "cardId", "account", "terminalId")):
                            rows_to_add.append((
                                mapping["customerId"],   # с префиксом
                                mapping["cardId"],      # с префиксом
                                mapping["account"],     # с префиксом
                                mapping["terminalId"]   # с префиксом
                            ))
                if rows_to_add:
                    append_rows_to_excel(xlsx_path, rows_to_add)

            progress_bar(total, total, prefix="Прогресс: ")
            sys.stdout.write("\n")

            # Форматирование Excel: убрать префиксы у всех 4 полей
            strip_prefixes_in_excel(xlsx_path)
            print("Обработка CSV-файлов для EXEL-таблицы завершена.")

            # Дедупликация Excel
            run_excel_dedup_with_notice(xlsx_path)

            break

        except KeyboardInterrupt:
            print("\nОстановлено пользователем.")
            break
        except RuntimeError as e:
            if str(e) == "OPENPYXL_MISSING":
                break
            raise

if __name__ == "__main__":
    main()
